import openpyxl

def ReadData(file,sheetname,rownum,colnum):
    workbook  = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=colnum).value

def WriteData(file,sheetname,rownum,colnum,data):
    workbook =openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum,column=colnum).value = data
    workbook.save(file)

def RowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row
